import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def analyze_gold_prices(excel_file):
    """
    엑셀 파일의 금 시세 데이터를 분석하고 통계값을 계산하여 엑셀에 기록합니다.
    """
    # 엑셀 파일 읽기
    print(f"엑셀 파일 읽는 중: {excel_file}")
    df = pd.read_excel(excel_file)
    
    print(f"\n데이터 정보:")
    print(f"총 행 수: {len(df)}")
    print(f"\n데이터 미리보기:")
    print(df.head())
    print(f"\n데이터 타입:")
    print(df.dtypes)
    
    # 숫자 데이터 전처리 (쉼표 제거 및 숫자 변환)
    numeric_columns = ['내가 살 때(3.75g) - 순금', '내가 팔 때(3.75g) - 순금', 
                      '내가 팔 때(3.75g) - 18K', '내가 팔 때(3.75g) - 14K']
    
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce')
    
    # 통계값 계산
    print("\n통계값 계산 중...")
    stats_data = []
    
    for col in numeric_columns:
        if col in df.columns:
            values = df[col].dropna()
            if len(values) > 0:
                stats_data.append({
                    '항목': col,
                    '평균': round(values.mean(), 2),
                    '최대값': round(values.max(), 2),
                    '최소값': round(values.min(), 2),
                    '중앙값': round(values.median(), 2),
                    '표준편차': round(values.std(), 2),
                    '데이터 개수': len(values)
                })
    
    stats_df = pd.DataFrame(stats_data)
    
    print("\n계산된 통계값:")
    print(stats_df)
    
    # 엑셀 파일에 통계값 추가
    print(f"\n엑셀 파일에 통계값 기록 중...")
    
    # openpyxl로 워크북 열기
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # 스타일 설정
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    stat_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 원본 데이터 다음 행에 통계 테이블 추가
    start_row = len(df) + 4  # 원본 데이터와 2행 간격
    
    # 제목 추가
    title_cell = ws.cell(row=start_row, column=1)
    title_cell.value = "통계 요약"
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f'A{start_row}:G{start_row}')
    title_cell.alignment = center_align
    
    # 헤더 작성
    headers = ['항목', '평균', '최대값', '최소값', '중앙값', '표준편차', '데이터 개수']
    header_row = start_row + 2
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    # 통계 데이터 작성
    for row_idx, stat_row in enumerate(stats_data, start=header_row + 1):
        col_idx = 1
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = stat_row['항목']
        cell.fill = stat_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        for stat_key in ['평균', '최대값', '최소값', '중앙값', '표준편차', '데이터 개수']:
            col_idx += 1
            cell = ws.cell(row=row_idx, column=col_idx)
            value = stat_row[stat_key]
            if isinstance(value, float):
                cell.value = round(value, 2)
                cell.number_format = '#,##0.00'
            else:
                cell.value = value
            cell.fill = stat_fill
            cell.border = border
            cell.alignment = center_align
    
    # 열 너비 자동 조정
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15
    
    # 추가 통계 정보 (요약)
    summary_row = row_idx + 3
    summary_cell = ws.cell(row=summary_row, column=1)
    summary_cell.value = "추가 분석"
    summary_cell.font = Font(bold=True, color="FFFFFF", size=12)
    summary_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{summary_row}:G{summary_row}')
    summary_cell.alignment = center_align
    
    # 가격 변동률 계산
    analysis_row = summary_row + 2
    if '내가 살 때(3.75g) - 순금' in df.columns:
        buy_prices = df['내가 살 때(3.75g) - 순금'].dropna()
        if len(buy_prices) > 1:
            price_change = buy_prices.iloc[0] - buy_prices.iloc[-1]
            price_change_pct = (price_change / buy_prices.iloc[-1]) * 100
            
            ws.cell(row=analysis_row, column=1, value="순금 구매가 변동률 (최신 vs 최초)")
            ws.cell(row=analysis_row, column=2, value=f"{price_change_pct:.2f}%")
            ws.cell(row=analysis_row, column=3, value=f"({price_change:,.0f}원)")
            
            analysis_row += 1
    
    if '내가 팔 때(3.75g) - 순금' in df.columns:
        sell_prices = df['내가 팔 때(3.75g) - 순금'].dropna()
        if len(sell_prices) > 1:
            price_change = sell_prices.iloc[0] - sell_prices.iloc[-1]
            price_change_pct = (price_change / sell_prices.iloc[-1]) * 100
            
            ws.cell(row=analysis_row, column=1, value="순금 판매가 변동률 (최신 vs 최초)")
            ws.cell(row=analysis_row, column=2, value=f"{price_change_pct:.2f}%")
            ws.cell(row=analysis_row, column=3, value=f"({price_change:,.0f}원)")
    
    # 파일 저장
    output_file = excel_file.replace('.xlsx', '_통계분석.xlsx')
    wb.save(output_file)
    
    print(f"\n통계값이 추가된 파일이 저장되었습니다: {output_file}")
    print(f"\n통계 요약:")
    print(stats_df.to_string(index=False))
    
    return output_file

if __name__ == "__main__":
    # 가장 최근 파일 사용
    import glob
    files = glob.glob("금시세_*.xlsx")
    if files:
        # 통계 분석 파일 제외
        files = [f for f in files if '_통계분석' not in f]
        if files:
            latest_file = max(files)
            analyze_gold_prices(latest_file)
        else:
            print("분석할 파일을 찾을 수 없습니다.")
    else:
        print("엑셀 파일을 찾을 수 없습니다.")

