# Yahoo Finance 주식 상승 종목 스크래핑
# -*- coding: utf-8 -*-

import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import re

def setup_driver():
    """Chrome 드라이버 설정"""
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # 헤드리스 모드
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

def parse_price_data(price_text):
    """가격 데이터 파싱"""
    # "6.96 +1.13 (+19.38%)" 형식에서 현재가 추출
    try:
        price_match = re.match(r'([\d,.]+)', price_text)
        if price_match:
            return float(price_match.group(1).replace(',', ''))
    except:
        pass
    return None

def parse_change(change_text):
    """변동값 파싱"""
    try:
        clean = change_text.replace('+', '').replace(',', '').strip()
        return float(clean)
    except:
        return None

def parse_percent(percent_text):
    """변동률 파싱"""
    try:
        clean = percent_text.replace('+', '').replace('%', '').replace(',', '').strip()
        return float(clean)
    except:
        return None

def parse_volume(vol_text):
    """거래량 파싱 (M=백만, B=십억, K=천)"""
    try:
        vol_text = vol_text.strip().upper()
        if 'B' in vol_text:
            return float(vol_text.replace('B', '').replace(',', '')) * 1_000_000_000
        elif 'M' in vol_text:
            return float(vol_text.replace('M', '').replace(',', '')) * 1_000_000
        elif 'K' in vol_text:
            return float(vol_text.replace('K', '').replace(',', '')) * 1_000
        else:
            return float(vol_text.replace(',', ''))
    except:
        return None

def parse_market_cap(cap_text):
    """시가총액 파싱"""
    return parse_volume(cap_text)  # 동일한 형식 사용

def parse_pe(pe_text):
    """P/E 비율 파싱"""
    try:
        if '--' in pe_text or pe_text.strip() == '':
            return None
        return float(pe_text.replace(',', '').strip())
    except:
        return None

def scrape_stock_gainers(target_count=50):
    """주식 상승 종목 스크래핑"""
    print("🚀 Yahoo Finance 주식 상승 종목 스크래핑 시작...")
    
    driver = setup_driver()
    stocks_data = []
    seen_symbols = set()  # 중복 방지
    
    try:
        url = "https://finance.yahoo.com/markets/stocks/gainers/"
        driver.get(url)
        
        # 페이지 로드 대기
        time.sleep(4)
        
        # 쿠키/팝업 닫기 시도
        try:
            consent_btn = driver.find_element(By.CSS_SELECTOR, "button.accept-all, button[name='agree']")
            consent_btn.click()
            time.sleep(1)
        except:
            pass
        
        # 테이블 로드 대기
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
        )
        
        collected = 0
        page = 1
        max_pages = 5  # 최대 5페이지까지 시도
        
        while collected < target_count and page <= max_pages:
            print(f"\n📄 페이지 {page} 스크래핑 중...")
            
            # 잠시 대기 (데이터 로드)
            time.sleep(2)
            
            # 테이블 행 찾기
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            
            if not rows:
                print("❌ 테이블 행을 찾을 수 없습니다.")
                break
            
            page_items = 0
            for row in rows:
                if collected >= target_count:
                    break
                    
                try:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) < 10:
                        continue
                    
                    # 각 셀에서 데이터 추출
                    symbol = cells[0].text.strip()
                    
                    # 중복 확인
                    if symbol in seen_symbols:
                        continue
                    
                    company_name = cells[1].text.strip()
                    
                    # 가격 정보 (현재가 + 변동)
                    price_cell = cells[3].text.strip()
                    current_price = parse_price_data(price_cell)
                    
                    # 변동값
                    change = parse_change(cells[4].text.strip())
                    
                    # 변동률
                    change_pct = parse_percent(cells[5].text.strip())
                    
                    # 거래량
                    volume = parse_volume(cells[6].text.strip())
                    
                    # 평균 거래량
                    avg_volume = parse_volume(cells[7].text.strip())
                    
                    # 시가총액
                    market_cap = parse_market_cap(cells[8].text.strip())
                    
                    # P/E 비율
                    pe_ratio = parse_pe(cells[9].text.strip())
                    
                    # 52주 변동률
                    ytd_change = None
                    if len(cells) > 10:
                        ytd_change = parse_percent(cells[10].text.strip())
                    
                    # 52주 범위
                    week52_range = ""
                    if len(cells) > 11:
                        week52_range = cells[11].text.strip()
                    
                    if symbol:  # 유효한 데이터만 추가
                        seen_symbols.add(symbol)
                        stock_info = {
                            'Symbol': symbol,
                            'Company Name': company_name,
                            'Price': current_price,
                            'Change': change,
                            'Change %': change_pct,
                            'Volume': volume,
                            'Avg Volume': avg_volume,
                            'Market Cap': market_cap,
                            'P/E Ratio': pe_ratio,
                            'YTD Change %': ytd_change,
                            '52 Week Range': week52_range
                        }
                        stocks_data.append(stock_info)
                        collected += 1
                        page_items += 1
                        print(f"  ✅ {collected}. {symbol}: {company_name} - ${current_price} ({change_pct}%)")
                        
                except Exception as e:
                    print(f"  ⚠️ 행 파싱 오류: {e}")
                    continue
            
            print(f"   현재 페이지에서 {page_items}개 수집")
            
            # 다음 페이지로 이동
            if collected < target_count:
                try:
                    # 여러 가지 방법으로 다음 버튼 찾기
                    next_btn = None
                    
                    # 방법 1: aria-label로 찾기
                    try:
                        next_btn = driver.find_element(By.CSS_SELECTOR, "button[aria-label='Goto next page']")
                    except:
                        pass
                    
                    # 방법 2: title로 찾기
                    if not next_btn:
                        try:
                            next_btn = driver.find_element(By.CSS_SELECTOR, "button[title*='next' i], button[title*='Next']")
                        except:
                            pass
                    
                    # 방법 3: 텍스트로 찾기
                    if not next_btn:
                        try:
                            buttons = driver.find_elements(By.TAG_NAME, "button")
                            for btn in buttons:
                                aria_label = btn.get_attribute("aria-label") or ""
                                if "next" in aria_label.lower():
                                    next_btn = btn
                                    break
                        except:
                            pass
                    
                    # 방법 4: SVG 화살표가 있는 버튼 찾기
                    if not next_btn:
                        try:
                            # 페이지네이션 영역 찾기
                            pagination = driver.find_element(By.CSS_SELECTOR, "div[class*='pagination'], nav[class*='pagination']")
                            buttons = pagination.find_elements(By.TAG_NAME, "button")
                            if len(buttons) >= 4:
                                next_btn = buttons[2]  # 보통 3번째 버튼이 next
                        except:
                            pass
                    
                    if next_btn:
                        # disabled 체크
                        is_disabled = next_btn.get_attribute("disabled")
                        if not is_disabled:
                            driver.execute_script("arguments[0].click();", next_btn)
                            time.sleep(3)
                            page += 1
                        else:
                            print("   다음 버튼이 비활성화 상태입니다.")
                            break
                    else:
                        print("   다음 버튼을 찾을 수 없습니다.")
                        break
                        
                except Exception as e:
                    print(f"⚠️ 다음 페이지 이동 오류: {e}")
                    break
        
    except Exception as e:
        print(f"❌ 스크래핑 오류: {e}")
        import traceback
        traceback.print_exc()
    finally:
        driver.quit()
    
    return stocks_data

def save_to_excel(data, filename=None):
    """데이터를 엑셀 파일로 저장"""
    if not data:
        print("❌ 저장할 데이터가 없습니다.")
        return None
    
    df = pd.DataFrame(data)
    
    # 52주 범위 데이터 정리 (줄바꿈을 공백으로 변경하고 저가/고가 분리)
    if '52 Week Range' in df.columns:
        df['52 Week Range'] = df['52 Week Range'].str.replace('\n', ' - ')
        # 52주 저가, 고가 분리
        try:
            df[['52 Week Low', '52 Week High']] = df['52 Week Range'].str.split(' - ', expand=True)
            df['52 Week Low'] = pd.to_numeric(df['52 Week Low'], errors='coerce')
            df['52 Week High'] = pd.to_numeric(df['52 Week High'], errors='coerce')
            df = df.drop(columns=['52 Week Range'])
        except:
            pass
    
    # 파일명 생성
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"주식상승종목_{timestamp}.xlsx"
    
    # 숫자 포맷팅을 위한 컬럼 타입 설정
    for col in ['Price', 'Change', 'Change %', 'Volume', 'Avg Volume', 'Market Cap', 'P/E Ratio', 'YTD Change %']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # 컬럼 순서 정리
    column_order = ['Symbol', 'Company Name', 'Price', 'Change', 'Change %', 
                    'Volume', 'Avg Volume', 'Market Cap', 'P/E Ratio', 
                    'YTD Change %', '52 Week Low', '52 Week High']
    df = df[[col for col in column_order if col in df.columns]]
    
    # 엑셀 파일로 저장
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Stock Gainers', index=False)
        
        # 워크시트 가져오기
        worksheet = writer.sheets['Stock Gainers']
        
        # 헤더 스타일 설정
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더 행 스타일 적용
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 데이터 셀 스타일
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 컬럼 너비 자동 조정
        column_widths = {
            'A': 10,   # Symbol
            'B': 35,   # Company Name
            'C': 12,   # Price
            'D': 10,   # Change
            'E': 12,   # Change %
            'F': 15,   # Volume
            'G': 15,   # Avg Volume
            'H': 15,   # Market Cap
            'I': 12,   # P/E Ratio
            'J': 15,   # YTD Change %
            'K': 15,   # 52 Week Low
            'L': 15,   # 52 Week High
        }
        
        for col_letter, width in column_widths.items():
            if col_letter in [c.column_letter for c in worksheet[1]]:
                worksheet.column_dimensions[col_letter].width = width
        
        # 첫 행 고정
        worksheet.freeze_panes = 'A2'
    
    print(f"\n✅ 엑셀 파일 저장 완료: {filename}")
    print(f"   총 {len(df)}개 종목 저장됨")
    
    # 요약 정보 출력
    print(f"\n📈 요약 정보:")
    print(f"   평균 상승률: {df['Change %'].mean():.2f}%")
    print(f"   최대 상승률: {df['Change %'].max():.2f}% ({df.loc[df['Change %'].idxmax(), 'Symbol']})")
    print(f"   최소 상승률: {df['Change %'].min():.2f}% ({df.loc[df['Change %'].idxmin(), 'Symbol']})")
    
    return filename

def main():
    """메인 실행 함수"""
    print("=" * 60)
    print("Yahoo Finance 주식 상승 종목 스크래핑")
    print("=" * 60)
    
    # 스크래핑 실행
    data = scrape_stock_gainers(target_count=50)
    
    if data:
        # 엑셀 저장
        filename = save_to_excel(data)
        
        # 데이터 미리보기
        print("\n📊 데이터 미리보기:")
        df = pd.DataFrame(data)
        print(df.head(10).to_string())
        
        return filename
    else:
        print("❌ 데이터를 가져오지 못했습니다.")
        return None

if __name__ == "__main__":
    main()

